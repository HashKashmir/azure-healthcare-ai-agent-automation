"""
Healthcare AI Agent — Web Dashboard
Run:  python dashboard.py
Open: http://localhost:8000
"""

import glob
import json
import os
import subprocess
import sys
from contextlib import asynccontextmanager

import uvicorn
from dotenv import load_dotenv
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse

load_dotenv()

BASE = os.path.dirname(os.path.abspath(__file__))

_api_procs: list = []


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup: launch mock APIs
    v = subprocess.Popen(
        [sys.executable, "-m", "uvicorn", "mock_apis.primehr_api:app", "--port", "8001"],
        cwd=BASE, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    m = subprocess.Popen(
        [sys.executable, "-m", "uvicorn", "mock_apis.claimbridge_api:app", "--port", "8002"],
        cwd=BASE, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    _api_procs.extend([v, m])
    yield
    # Shutdown: stop mock APIs
    for p in _api_procs:
        try:
            p.terminate()
        except Exception:
            pass


app = FastAPI(lifespan=lifespan)


# ── SSE helpers ───────────────────────────────────────────────────────────────

SSE_HEADERS = {"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}


def _stream(cmd: list):
    """Run a subprocess and stream its output as SSE."""
    def generate():
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", cwd=BASE,
        )
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                yield f"data: {json.dumps(line)}\n\n"
        proc.wait()
        yield f"data: {json.dumps('__DONE__')}\n\n"
    return StreamingResponse(generate(), media_type="text/event-stream", headers=SSE_HEADERS)


# ── Automation endpoints ──────────────────────────────────────────────────────

@app.get("/run/onboarding")
def run_onboarding(employee_id: str = "EMP-0023"):
    return _stream([sys.executable, "-m", "agent.agent_runner",
                    "--automation", "onboarding", "--employee-id", employee_id])


@app.get("/run/billing")
def run_billing():
    return _stream([sys.executable, "-m", "agent.agent_runner", "--automation", "billing"])


@app.get("/run/report")
def run_report(mode: str = "financial"):
    if mode == "all":
        def generate():
            for m in ["financial", "clinical", "billing"]:
                yield f"data: {json.dumps(f'[dashboard] Starting {m} report...')}\n\n"
                proc = subprocess.Popen(
                    [sys.executable, "-m", "agent.agent_runner",
                     "--automation", "report", "--mode", m],
                    stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace", cwd=BASE,
                )
                for line in proc.stdout:
                    line = line.rstrip()
                    if line:
                        yield f"data: {json.dumps(line)}\n\n"
                proc.wait()
            yield f"data: {json.dumps('[dashboard] Building combined tabbed report...')}\n\n"
            try:
                import sys as _sys
                _sys.path.insert(0, BASE)
                from reporting.report_combiner import build_combined_report
                result = build_combined_report()
                if result.get("success"):
                    yield f"data: {json.dumps('[dashboard] Combined report ready — click View Report.')}\n\n"
                else:
                    yield f"data: {json.dumps('[dashboard] Combine step: ' + result.get('error','unknown error'))}\n\n"
            except Exception as _e:
                yield f"data: {json.dumps('[dashboard] Combine step skipped: ' + str(_e))}\n\n"
            yield f"data: {json.dumps('__DONE__')}\n\n"
        return StreamingResponse(generate(), media_type="text/event-stream", headers=SSE_HEADERS)
    return _stream([sys.executable, "-m", "agent.agent_runner",
                    "--automation", "report", "--mode", mode])


@app.post("/upload/bulk-onboarding")
async def upload_bulk_onboarding(file: UploadFile = File(...)):
    try:
        import io
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        employee_ids = []
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if first_row:
                first_row = False
                if row and row[0] and str(row[0]).strip().lower() in ("employee_id", "id", "emp_id"):
                    continue
            if row and row[0]:
                val = str(row[0]).strip().upper()
                if val.startswith("EMP-"):
                    employee_ids.append(val)
        wb.close()
        return {"employee_ids": employee_ids, "count": len(employee_ids)}
    except Exception as exc:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {exc}")


@app.get("/bulk-onboarding/template")
def download_bulk_template():
    path = os.path.join(BASE, "data", "bulk_onboarding_template.xlsx")
    if not os.path.exists(path):
        return HTMLResponse("<p>Template not found.</p>", status_code=404)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_onboarding_template.xlsx"},
    )


@app.get("/run/bulk-onboarding")
def run_bulk_onboarding(ids: str = ""):
    if not ids.strip():
        def _empty():
            yield f"data: {json.dumps('[ERROR] No employee IDs provided.')}\n\n"
            yield f"data: {json.dumps('__DONE__')}\n\n"
        return StreamingResponse(_empty(), media_type="text/event-stream", headers=SSE_HEADERS)
    return _stream([sys.executable, "-m", "agent.bulk_onboarding_runner", "--employee-ids", ids])


@app.post("/upload/bulk-intake")
async def upload_bulk_intake(file: UploadFile = File(...)):
    try:
        import io
        import openpyxl
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        headers = None
        patients = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c).strip().lower() if c else "" for c in row]
                continue
            if not any(row):
                continue
            p = {headers[i]: (str(row[i]).strip() if row[i] is not None else "")
                 for i in range(min(len(headers), len(row)))}
            if p.get("first_name") or p.get("last_name"):
                patients.append(p)
        wb.close()
        patients_path = os.path.join(BASE, "data", "bulk_intake_patients.json")
        with open(patients_path, "w", encoding="utf-8") as f:
            json.dump(patients, f)
        return {"patients": patients, "count": len(patients)}
    except Exception as exc:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {exc}")


@app.get("/bulk-intake/template")
def download_bulk_intake_template():
    path = os.path.join(BASE, "data", "bulk_intake_template.xlsx")
    if not os.path.exists(path):
        return HTMLResponse("<p>Template not found.</p>", status_code=404)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_intake_template.xlsx"},
    )


@app.get("/run/bulk-intake")
def run_bulk_intake():
    patients_path = os.path.join(BASE, "data", "bulk_intake_patients.json")
    if not os.path.exists(patients_path):
        def _empty():
            yield f"data: {json.dumps('[ERROR] No patients file found. Upload an Excel file first.')}\n\n"
            yield f"data: {json.dumps('__DONE__')}\n\n"
        return StreamingResponse(_empty(), media_type="text/event-stream", headers=SSE_HEADERS)
    return _stream([sys.executable, "-m", "agent.bulk_intake_runner",
                    "--patients-file", patients_path])


_NO_CACHE = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}


@app.get("/onboarding/pdf/{employee_id}")
def view_onboarding_pdf(employee_id: str):
    docs_dir = os.path.join(BASE, "data", "onboarding_docs")
    files = sorted(
        glob.glob(os.path.join(docs_dir, f"{employee_id.upper()}_*.pdf")),
        key=os.path.getmtime, reverse=True,
    )
    if files:
        return FileResponse(files[0], media_type="application/pdf", headers=_NO_CACHE)
    return HTMLResponse("<p>No onboarding PDF found for this employee yet.</p>", status_code=404)


@app.get("/intake/bulk-pdf/{key}")
def view_bulk_intake_pdf(key: str):
    pdf_path = os.path.join(BASE, "data", f"intake_bulk_{key}.pdf")
    if os.path.exists(pdf_path):
        return FileResponse(pdf_path, media_type="application/pdf", headers=_NO_CACHE)
    return HTMLResponse("<p>PDF not found.</p>", status_code=404)


@app.get("/run/intake")
def run_intake():
    def generate():
        yield f"data: {json.dumps('[dashboard] Generating patient intake PDF...')}\n\n"
        proc = subprocess.Popen(
            [sys.executable, "scripts/generate_intake_pdf.py"],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", cwd=BASE,
        )
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                yield f"data: {json.dumps(line)}\n\n"
        proc.wait()

        url_file = os.path.join(BASE, "data", "intake_blob_url.txt")
        try:
            with open(url_file) as f:
                blob_url = f.read().strip()
        except FileNotFoundError:
            yield f"data: {json.dumps('[ERROR] Could not read blob URL.')}\n\n"
            yield f"data: {json.dumps('__DONE__')}\n\n"
            return

        yield f"data: {json.dumps('[dashboard] Running intake agent...')}\n\n"
        proc = subprocess.Popen(
            [sys.executable, "-m", "agent.agent_runner",
             "--automation", "intake", "--blob-url", blob_url],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", cwd=BASE,
        )
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                yield f"data: {json.dumps(line)}\n\n"
        proc.wait()
        yield f"data: {json.dumps('__DONE__')}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream", headers=SSE_HEADERS)


# ── Data endpoints ────────────────────────────────────────────────────────────

@app.get("/api/status")
def api_status():
    import requests as req
    results = {}
    for name, port in [("PrimeHR", 8001), ("ClaimBridge", 8002)]:
        try:
            req.get(f"http://localhost:{port}/health", timeout=1)
            results[name] = "ready"
        except Exception:
            results[name] = "starting"
    return results


@app.get("/audit")
def get_audit():
    log_path = os.path.join(BASE, "audit", "audit_log.jsonl")
    entries = []
    try:
        with open(log_path, encoding="utf-8") as f:
            lines = f.readlines()
        for line in lines[-60:]:
            try:
                entries.append(json.loads(line))
            except Exception:
                pass
    except FileNotFoundError:
        pass
    return {"entries": list(reversed(entries))}


@app.get("/billing/report")
def view_billing_report():
    reports_dir = os.path.join(BASE, "data", "reports")
    files = sorted(glob.glob(os.path.join(reports_dir, "billing_triage_*.html")),
                   key=os.path.getmtime, reverse=True)
    if files:
        return FileResponse(files[0], media_type="text/html")
    return HTMLResponse("<p>No billing report yet. Run Automation 3 first.</p>")


@app.get("/report/view")
def view_report():
    reports_dir = os.path.join(BASE, "data", "reports")
    files = sorted(glob.glob(os.path.join(reports_dir, "RPT-*.html")),
                   key=os.path.getmtime, reverse=True)
    if files:
        return FileResponse(files[0], media_type="text/html")
    return HTMLResponse("<p>No analytics report yet. Run Automation 4 first.</p>")


@app.get("/intake/pdf")
def view_pdf():
    pdf_path = os.path.join(BASE, "data", "sample_intake.pdf")
    if os.path.exists(pdf_path):
        return FileResponse(pdf_path, media_type="application/pdf", headers=_NO_CACHE)
    return HTMLResponse("<p>No PDF found yet. Run Automation 2 first.</p>")


# ── Dashboard UI ──────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
def index():
    return DASHBOARD_HTML


DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Healthcare AI Dashboard</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#F0F4F8;color:#212121}
.header{background:#1565C0;color:#fff;padding:16px 32px;display:flex;align-items:center;gap:16px;box-shadow:0 2px 8px rgba(0,0,0,.2)}
.header h1{font-size:20px;font-weight:700}
.header .sub{font-size:12px;opacity:.75;margin-top:2px}
.api-badge{margin-left:auto;display:flex;align-items:center;gap:8px;background:rgba(255,255,255,.15);padding:6px 14px;border-radius:20px}
.dot{width:9px;height:9px;border-radius:50%;background:#FF9800}
.dot.ready{background:#4CAF50}
.api-label{font-size:12px}
.container{max-width:1120px;margin:0 auto;padding:24px}
.section-title{font-size:13px;font-weight:700;color:#9E9E9E;text-transform:uppercase;letter-spacing:.8px;margin-bottom:14px}
.cards{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-bottom:22px}
.card{background:#fff;border-radius:10px;padding:22px;box-shadow:0 2px 10px rgba(0,0,0,.07);border-top:4px solid #1565C0;display:flex;flex-direction:column}
.card-head{display:flex;align-items:flex-start;gap:12px;margin-bottom:12px}
.icon{font-size:26px;line-height:1;flex-shrink:0}
.card-title{font-size:14px;font-weight:700;color:#1565C0}
.card-sub{font-size:11px;color:#9E9E9E;margin-top:2px}
.card-desc{font-size:13px;color:#424242;line-height:1.6;margin-bottom:10px;flex:1}
.card-tools{font-size:11px;color:#BDBDBD;font-family:monospace;margin-bottom:14px;line-height:1.7}
.controls{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.badge{display:inline-block;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:700}
.badge-idle{background:#E3F2FD;color:#1565C0}
.badge-running{background:#FFF3E0;color:#E65100;animation:pulse 1.4s infinite}
.badge-success{background:#E8F5E9;color:#2E7D32}
.badge-error{background:#FFEBEE;color:#C62828}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.4}}
.btn{padding:7px 18px;border:none;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;transition:background .15s}
.btn-primary{background:#1565C0;color:#fff}
.btn-primary:hover{background:#0D47A1}
.btn-primary:disabled{background:#90CAF9;cursor:not-allowed}
.btn-ghost{background:#E3F2FD;color:#1565C0;padding:7px 14px}
.btn-ghost:hover{background:#BBDEFB}
select,input{padding:6px 10px;border:1px solid #BBDEFB;border-radius:6px;font-size:12px;color:#1565C0;outline:none}
input{width:110px}
.result-row{margin-top:12px;padding-top:12px;border-top:1px solid #F0F0F0;font-size:12px;display:none;align-items:center;gap:10px;flex-wrap:wrap}
.result-row.show{display:flex}
.result-row .ok{color:#2E7D32;font-weight:600}
.result-row a{color:#1565C0;text-decoration:none;font-weight:600}
.result-row a:hover{text-decoration:underline}
.console-wrap{background:#fff;border-radius:10px;box-shadow:0 2px 10px rgba(0,0,0,.07);overflow:hidden;margin-bottom:22px}
.console-top{background:#263238;color:#90A4AE;padding:10px 16px;font-size:12px;font-weight:600;display:flex;justify-content:space-between;align-items:center}
.console-top button{background:none;border:1px solid #546E7A;color:#90A4AE;padding:3px 10px;border-radius:4px;cursor:pointer;font-size:11px}
.console-top button:hover{background:#37474F}
#console{background:#1E1E1E;padding:14px 16px;height:260px;overflow-y:auto;font-family:'Courier New',monospace;font-size:12px;line-height:1.75;color:#CCC}
.ph{color:#546E7A;font-style:italic}
.lt{color:#80CBC4}
.lr{color:#C3E88D}
.li{color:#FFCB6B}
.lf{color:#82AAFF;font-weight:700}
.ld{color:#F48FB1}
.audit-wrap{background:#fff;border-radius:10px;box-shadow:0 2px 10px rgba(0,0,0,.07);overflow:hidden}
.audit-top{background:#1565C0;color:#fff;padding:12px 20px;font-size:14px;font-weight:600;display:flex;justify-content:space-between;align-items:center}
.audit-top button{background:rgba(255,255,255,.2);border:none;color:#fff;padding:4px 12px;border-radius:4px;cursor:pointer;font-size:12px}
.audit-cols{padding:8px 20px;display:grid;grid-template-columns:150px 90px 160px 1fr;gap:10px;font-size:11px;font-weight:700;color:#9E9E9E;text-transform:uppercase;letter-spacing:.5px;border-bottom:2px solid #E0E0E0}
#audit-body{max-height:280px;overflow-y:auto}
.audit-row{padding:9px 20px;display:grid;grid-template-columns:150px 90px 160px 1fr;gap:10px;font-size:12px;border-bottom:1px solid #F5F5F5;align-items:center}
.audit-row:hover{background:#F5F7FA}
.at{color:#9E9E9E;font-family:monospace;font-size:11px}
.aa{font-weight:700;color:#1565C0;text-transform:capitalize}
.ac{font-family:monospace;font-size:11px;color:#37474F}
.ae{color:#757575}
.empty{padding:20px;color:#9E9E9E;font-size:13px;text-align:center}
.rmode-btn{padding:5px 16px;border:1px solid #BBDEFB;border-radius:16px;background:#fff;color:#1565C0;font-size:12px;font-weight:600;cursor:pointer;transition:all .15s}
.rmode-btn:hover{background:#E3F2FD}
.rmode-btn.rmode-active{background:#1565C0;color:#fff;border-color:#1565C0}
</style>
</head>
<body>

<div class="header">
  <div>
    <div style="font-size:10px;letter-spacing:1.5px;text-transform:uppercase;opacity:.7;margin-bottom:4px">Healthcare AI</div>
    <h1>Admin Agent Dashboard</h1>
    <div class="sub">Four automated workflows — click Run to execute</div>
  </div>
  <div class="api-badge">
    <div class="dot" id="api-dot"></div>
    <span class="api-label" id="api-label">Checking APIs...</span>
  </div>
</div>

<div class="container">

  <div class="section-title">Automations</div>

  <div class="cards">

    <!-- Card 1: Onboarding -->
    <div class="card">
      <div class="card-head">
        <div class="icon">&#128100;</div>
        <div>
          <div class="card-title">Automation 1 &mdash; Employee Onboarding</div>
          <div class="card-sub">PrimeHR HR System &rarr; Azure Blob &rarr; Email</div>
        </div>
      </div>
      <div class="card-desc">
        Retrieves the employee record from PrimeHR, selects role-appropriate forms (clinical staff get license verification; physicians get credentialing forms; contractors get W-9 instead of W-4), assesses compliance risk based on pending documents and start date, generates an onboarding PDF, emails the new hire their document checklist, and sends the manager a standard, conditional, or urgent notification based on risk level.
      </div>
      <div class="card-tools">get_employee_details &middot; fill_onboarding_form &middot; assess_onboarding_risk &middot; store_document &middot; notify_employee &middot; notify_manager</div>
      <div class="controls">
        <span class="badge badge-idle" id="badge-onboarding">Idle</span>
        <input id="emp-id" value="EMP-0023" title="Employee ID (EMP-0001 to EMP-0075)"/>
        <button class="btn btn-primary" id="btn-onboarding" onclick="run('onboarding')">&#9654; Run Single</button>
        <a id="link-onboarding-pdf" href="#" target="_blank" style="display:none"><button class="btn btn-ghost">View PDF</button></a>
      </div>
      <div class="result-row" id="res-onboarding">
        <span class="ok">&#10003; Complete</span>
        <span style="color:#555">Check Gmail for the manager notification email.</span>
      </div>
      <div style="border-top:1px dashed #BBDEFB;margin-top:14px;padding-top:14px">
        <div style="font-size:11px;color:#9E9E9E;font-weight:700;letter-spacing:.5px;margin-bottom:8px">BULK UPLOAD (.xlsx)</div>
        <div class="controls">
          <span class="badge badge-idle" id="badge-bulk">Idle</span>
          <input type="file" id="bulk-file-input" accept=".xlsx" style="display:none" onchange="onBulkFileSelect(this)"/>
          <button class="btn btn-ghost" onclick="document.getElementById('bulk-file-input').click()">&#128196; Choose File</button>
          <button class="btn btn-ghost" onclick="window.open('/bulk-onboarding/template','_blank')">&#8595; Template</button>
          <button class="btn btn-primary" id="btn-bulk" onclick="runBulk()" disabled>&#9654; Run Bulk</button>
        </div>
        <div id="bulk-file-info" style="font-size:11px;color:#9E9E9E;margin-top:6px">No file selected — upload .xlsx with employee_id column</div>
      </div>
    </div>

    <!-- Card 2: Intake -->
    <div class="card">
      <div class="card-head">
        <div class="icon">&#128203;</div>
        <div>
          <div class="card-title">Automation 2 &mdash; Patient Intake</div>
          <div class="card-sub">Azure Document Intelligence &rarr; Eligibility &rarr; DocVault</div>
        </div>
      </div>
      <div class="card-desc">
        Generates a random synthetic patient intake PDF, uploads it to Azure Blob, extracts structured fields via Document Intelligence, validates insurance eligibility, stores the indexed patient record, emails the patient a confirmation, and alerts the front desk if insurance is ineligible.
      </div>
      <div class="card-tools">extract_document_fields &middot; validate_insurance &middot; store_indexed_record &middot; notify_patient &middot; notify_staff_ineligible</div>
      <div class="controls">
        <span class="badge badge-idle" id="badge-intake">Idle</span>
        <button class="btn btn-primary" id="btn-intake" onclick="run('intake')">&#9654; Run Single</button>
        <button class="btn btn-ghost" onclick="window.open('/intake/pdf?t='+Date.now(),'_blank')">View PDF</button>
      </div>
      <div class="result-row" id="res-intake">
        <span class="ok">&#10003; Complete</span>
        <a href="/intake/pdf" target="_blank">View intake PDF &rarr;</a>
      </div>
      <div style="border-top:1px dashed #BBDEFB;margin-top:14px;padding-top:14px">
        <div style="font-size:11px;color:#9E9E9E;font-weight:700;letter-spacing:.5px;margin-bottom:8px">BULK UPLOAD (.xlsx)</div>
        <div class="controls">
          <span class="badge badge-idle" id="badge-bulk-intake">Idle</span>
          <input type="file" id="bulk-intake-file-input" accept=".xlsx" style="display:none" onchange="onBulkIntakeFileSelect(this)"/>
          <button class="btn btn-ghost" onclick="document.getElementById('bulk-intake-file-input').click()">&#128196; Choose File</button>
          <button class="btn btn-ghost" onclick="window.open('/bulk-intake/template','_blank')">&#8595; Template</button>
          <button class="btn btn-primary" id="btn-bulk-intake" onclick="runBulkIntake()" disabled>&#9654; Run Bulk</button>
        </div>
        <div id="bulk-intake-file-info" style="font-size:11px;color:#9E9E9E;margin-top:6px">No file selected — upload .xlsx with patient columns</div>
      </div>
    </div>

    <!-- Card 3: Billing -->
    <div class="card">
      <div class="card-head">
        <div class="icon">&#127973;</div>
        <div>
          <div class="card-title">Automation 3 &mdash; Billing Triage</div>
          <div class="card-sub">ClaimBridge &rarr; NLM ICD-10 API &rarr; Report + Emails</div>
        </div>
      </div>
      <div class="card-desc">
        Pulls the top 5 exception claims from the billing queue, validates each ICD-10 code against the live NLM government API, classifies rejection reasons, auto-resubmits what it can, emails staff alerts for escalated claims, generates a visual triage report, and emails the manager a financial summary.
      </div>
      <div class="card-tools">get_billing_exception_queue &middot; get_claim_details &middot; validate_icd10_code &middot; classify_claim &middot; resubmit_claim &middot; route_to_staff &middot; record_claim_outcome &middot; notify_staff_claim_routed &middot; generate_billing_report &middot; send_billing_summary_email</div>
      <div class="controls">
        <span class="badge badge-idle" id="badge-billing">Idle</span>
        <button class="btn btn-primary" id="btn-billing" onclick="run('billing')">&#9654; Run</button>
        <button class="btn btn-ghost" id="btn-view-billing-report" onclick="window.open('/billing/report','_blank')" style="display:none">View Report</button>
      </div>
      <div class="result-row" id="res-billing">
        <span class="ok">&#10003; Complete</span>
        <a href="/billing/report" target="_blank">View triage report &rarr;</a>
        <span style="color:#555">+ check Gmail for staff alerts and manager summary.</span>
      </div>
    </div>

    <!-- Card 4: Reporting -->
    <div class="card">
      <div class="card-head">
        <div class="icon">&#128202;</div>
        <div>
          <div class="card-title">Automation 4 &mdash; Weekly Report</div>
          <div class="card-sub">pandas &rarr; matplotlib &rarr; Azure Communication Services</div>
        </div>
      </div>
      <div class="card-desc">
        Analyzes 4 weeks of operational metrics across three report modes — Financial (CFO), Clinical (Operations), and Billing (Billing Manager). Pre-computes signals in pandas, sends a structured brief to o4-mini for interpretation, generates benchmark-anchored charts, assembles a styled HTML report, uploads it to Azure Blob, and emails it with inline charts to the admin list.
      </div>
      <div class="card-tools">fetch_data_csv &middot; analyze_trends &middot; generate_charts &middot; build_report &middot; send_report_email</div>
      <div class="controls">
        <span class="badge badge-idle" id="badge-report">Idle</span>
        <select id="report-mode">
          <option value="financial">Financial</option>
          <option value="clinical">Clinical</option>
          <option value="billing">Billing</option>
          <option value="all">All Modes</option>
        </select>
        <button class="btn btn-primary" id="btn-report" onclick="run('report')">&#9654; Run</button>
        <button class="btn btn-ghost" id="btn-view-report" onclick="window.open('/report/view','_blank')" style="display:none">View Report</button>
      </div>
      <div class="result-row" id="res-report">
        <span class="ok">&#10003; Complete</span>
        <a href="/report/view" target="_blank">Open report in browser &rarr;</a>
        <span style="color:#555">+ check Gmail for the email with charts.</span>
      </div>
    </div>

  </div>

  <!-- Bulk Onboarding Board -->
  <div id="bulk-board-wrap" style="display:none;margin-bottom:22px">
    <div class="section-title">Bulk Onboarding Board</div>
    <div style="background:#fff;border-radius:10px;padding:18px 24px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:14px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px">
      <div style="text-align:center">
        <div id="bko-total" style="font-size:28px;font-weight:700;color:#1565C0">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Total</div>
      </div>
      <div style="text-align:center">
        <div id="bko-complete" style="font-size:28px;font-weight:700;color:#2E7D32">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Complete</div>
      </div>
      <div style="text-align:center">
        <div id="bko-running" style="font-size:28px;font-weight:700;color:#E65100">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Running</div>
      </div>
      <div style="text-align:center">
        <div id="bko-failed" style="font-size:28px;font-weight:700;color:#C62828">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Failed</div>
      </div>
    </div>
    <div id="bulk-cards-grid"></div>
  </div>

  <!-- Bulk Intake Board -->
  <div id="bulk-intake-board-wrap" style="display:none;margin-bottom:22px">
    <div class="section-title">Bulk Intake Board</div>
    <div style="background:#fff;border-radius:10px;padding:18px 24px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:14px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px">
      <div style="text-align:center">
        <div id="bki-total" style="font-size:28px;font-weight:700;color:#1565C0">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Total</div>
      </div>
      <div style="text-align:center">
        <div id="bki-complete" style="font-size:28px;font-weight:700;color:#2E7D32">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Complete</div>
      </div>
      <div style="text-align:center">
        <div id="bki-eligible" style="font-size:28px;font-weight:700;color:#2E7D32">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Eligible</div>
      </div>
      <div style="text-align:center">
        <div id="bki-ineligible" style="font-size:28px;font-weight:700;color:#C62828">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Ineligible</div>
      </div>
    </div>
    <div id="bulk-intake-cards-grid"></div>
  </div>

  <!-- Billing Triage Board -->
  <div id="billing-board-wrap" style="display:none;margin-bottom:22px">
    <div class="section-title">Billing Triage Board</div>
    <div style="background:#fff;border-radius:10px;padding:18px 24px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:14px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px">
      <div style="text-align:center">
        <div id="bb-total" style="font-size:28px;font-weight:700;color:#1565C0">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Processed</div>
      </div>
      <div style="text-align:center">
        <div id="bb-resolved" style="font-size:28px;font-weight:700;color:#2E7D32">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Auto-Resolved</div>
      </div>
      <div style="text-align:center">
        <div id="bb-dollars-rec" style="font-size:28px;font-weight:700;color:#2E7D32">$0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Recovered</div>
      </div>
      <div style="text-align:center">
        <div id="bb-dollars-risk" style="font-size:28px;font-weight:700;color:#C62828">$0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">At Risk</div>
      </div>
    </div>
    <div id="billing-cards-grid"></div>
  </div>

  <!-- Report Insights Board -->
  <div id="report-board-wrap" style="display:none;margin-bottom:22px">
    <div class="section-title">Report Insights Board</div>
    <div style="background:#fff;border-radius:10px;padding:18px 24px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:14px;display:grid;grid-template-columns:repeat(4,1fr);gap:14px">
      <div style="text-align:center">
        <div id="rb-total" style="font-size:28px;font-weight:700;color:#1565C0">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Insights</div>
      </div>
      <div style="text-align:center">
        <div id="rb-critical" style="font-size:28px;font-weight:700;color:#C62828">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Critical</div>
      </div>
      <div style="text-align:center">
        <div id="rb-warning" style="font-size:28px;font-weight:700;color:#E65100">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">Warning</div>
      </div>
      <div style="text-align:center">
        <div id="rb-ok" style="font-size:28px;font-weight:700;color:#2E7D32">0</div>
        <div style="font-size:10px;color:#9E9E9E;text-transform:uppercase;letter-spacing:.6px;margin-top:4px">OK / Info</div>
      </div>
    </div>
    <div id="report-mode-tabs" style="display:none;background:#fff;border-radius:10px;padding:10px 16px;box-shadow:0 2px 10px rgba(0,0,0,.07);margin-bottom:10px;display:flex;gap:6px;flex-wrap:wrap">
      <button class="rmode-btn rmode-active" onclick="filterInsights('all')" id="rmode-all">All</button>
      <button class="rmode-btn" onclick="filterInsights('financial')" id="rmode-financial">Financial</button>
      <button class="rmode-btn" onclick="filterInsights('clinical')" id="rmode-clinical">Clinical</button>
      <button class="rmode-btn" onclick="filterInsights('billing')" id="rmode-billing">Billing</button>
    </div>
    <div id="report-insights-grid"></div>
  </div>

  <!-- Console -->
  <div class="section-title">Agent Output</div>
  <div class="console-wrap">
    <div class="console-top">
      <span id="console-label">Waiting for automation to run...</span>
      <button onclick="clearConsole()">Clear</button>
    </div>
    <div id="console"><span class="ph">Output will stream here in real time when you click Run above.</span></div>
  </div>

  <!-- Audit Log -->
  <div class="section-title">Audit Trail</div>
  <div class="audit-wrap">
    <div class="audit-top">
      All Tool Calls &mdash; Every Action Logged
      <button onclick="loadAudit()">&#8635; Refresh</button>
    </div>
    <div class="audit-cols">
      <span>Timestamp</span><span>Automation</span><span>Tool Called</span><span>Entity</span>
    </div>
    <div id="audit-body"><div class="empty">Loading...</div></div>
  </div>

</div>

<script>
function setBadge(id, state){
  const b = document.getElementById('badge-'+id);
  const labels = {idle:'Idle',running:'Running...',success:'Success',error:'Error'};
  b.className = 'badge badge-'+state;
  b.textContent = labels[state] || state;
}
function setBtn(id, disabled){
  const b = document.getElementById('btn-'+id);
  if(b) b.disabled = disabled;
}
function showResult(id){
  const el = document.getElementById('res-'+id);
  if(el) el.classList.add('show');
  if(id==='report') document.getElementById('btn-view-report').style.display='inline-block';
  if(id==='billing') document.getElementById('btn-view-billing-report').style.display='inline-block';
  if(id==='onboarding'){
    var pdfLink = document.getElementById('link-onboarding-pdf');
    var empId = encodeURIComponent(document.getElementById('emp-id').value||'EMP-0023');
    pdfLink.href = '/onboarding/pdf/'+empId+'?t='+Date.now();
    pdfLink.style.display = 'inline-block';
  }
  if(id==='intake'){
    var intakeLink = document.querySelector('#res-intake a[href*="intake/pdf"]');
    if(intakeLink) intakeLink.href = '/intake/pdf?t='+Date.now();
  }
}
var _billingOutcomes = [];
var _reportInsights  = [];
function clearReportBoard(){
  _reportInsights = [];
  document.getElementById('report-insights-grid').innerHTML = '';
  document.getElementById('report-board-wrap').style.display = 'none';
  document.getElementById('report-mode-tabs').style.display = 'none';
  document.getElementById('rb-total').textContent    = '0';
  document.getElementById('rb-critical').textContent = '0';
  document.getElementById('rb-warning').textContent  = '0';
  document.getElementById('rb-ok').textContent       = '0';
  document.querySelectorAll('.rmode-btn').forEach(function(b){ b.classList.remove('rmode-active'); });
  document.getElementById('rmode-all').classList.add('rmode-active');
}
function filterInsights(mode){
  document.querySelectorAll('.rmode-btn').forEach(function(b){ b.classList.remove('rmode-active'); });
  document.getElementById('rmode-'+mode).classList.add('rmode-active');
  document.querySelectorAll('.insight-card').forEach(function(c){
    c.style.display = (mode==='all' || c.dataset.mode===mode) ? 'block' : 'none';
  });
}
function _updateReportCounts(){
  var crit = _reportInsights.filter(function(x){ return x.severity==='critical'; }).length;
  var warn = _reportInsights.filter(function(x){ return x.severity==='warning'; }).length;
  var ok   = _reportInsights.filter(function(x){ return x.severity==='ok'||x.severity==='info'; }).length;
  document.getElementById('rb-total').textContent    = _reportInsights.length;
  document.getElementById('rb-critical').textContent = crit;
  document.getElementById('rb-warning').textContent  = warn;
  document.getElementById('rb-ok').textContent       = ok;
}
function addInsightCard(ins){
  _reportInsights.push(ins);
  document.getElementById('report-board-wrap').style.display = 'block';
  var mode = ins._mode || 'unknown';
  var sev  = ins.severity || 'info';
  var cfg  = {
    critical: {border:'#C62828', bg:'#FFEBEE', clr:'#C62828', icon:'&#9888;'},
    warning:  {border:'#E65100', bg:'#FFF3E0', clr:'#E65100', icon:'&#9888;'},
    ok:       {border:'#2E7D32', bg:'#E8F5E9', clr:'#2E7D32', icon:'&#10003;'},
    info:     {border:'#1565C0', bg:'#E3F2FD', clr:'#1565C0', icon:'&#9670;'},
  }[sev] || {border:'#9E9E9E', bg:'#F5F5F5', clr:'#616161', icon:'&#9670;'};
  var modeLabel = {financial:'Financial',clinical:'Clinical',billing:'Billing'}[mode] || mode;
  var div = document.createElement('div');
  div.className   = 'insight-card';
  div.dataset.mode = mode;
  div.style.cssText = 'background:#fff;border-radius:8px;border-left:5px solid '+cfg.border+';padding:16px 20px;box-shadow:0 1px 6px rgba(0,0,0,.07);margin-bottom:10px;opacity:0;transition:opacity .35s ease';
  div.innerHTML =
    '<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px">'
    +'<div style="font-weight:700;font-size:13px;color:'+cfg.clr+'">'+cfg.icon+'&nbsp;&nbsp;'+(ins.title||'')+'</div>'
    +'<div style="display:flex;gap:6px;align-items:center;flex-shrink:0;margin-left:12px">'
    +'<span style="background:#E3F2FD;color:#1565C0;font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px">'+modeLabel+'</span>'
    +'<span style="background:'+cfg.bg+';color:'+cfg.clr+';font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px">'+sev.toUpperCase()+'</span>'
    +'</div></div>'
    +'<div style="font-size:12px;color:#424242;line-height:1.6;margin-bottom:8px">'+(ins.detail||'')+'</div>'
    +'<div style="font-size:11px;color:#555"><strong>Recommendation:</strong> '+(ins.recommendation||'')+'</div>';
  document.getElementById('report-insights-grid').appendChild(div);
  setTimeout(function(){ div.style.opacity='1'; }, 40);
  // Show tab bar once more than one mode has insights
  var modes = [...new Set(_reportInsights.map(function(x){ return x._mode; }).filter(Boolean))];
  if(modes.length > 1) document.getElementById('report-mode-tabs').style.display = 'flex';
  _updateReportCounts();
}
function clearBillingBoard(){
  _billingOutcomes = [];
  document.getElementById('billing-cards-grid').innerHTML = '';
  document.getElementById('billing-board-wrap').style.display = 'none';
  document.getElementById('bb-total').textContent = '0';
  document.getElementById('bb-resolved').textContent = '0';
  document.getElementById('bb-dollars-rec').textContent = '$0';
  document.getElementById('bb-dollars-risk').textContent = '$0';
}
function addClaimCard(c){
  _billingOutcomes.push(c);
  document.getElementById('billing-board-wrap').style.display = 'block';
  var isRes = c.outcome === 'auto_resolved';
  var border = isRes ? '#2E7D32' : '#C62828';
  var bbg    = isRes ? '#E8F5E9' : '#FFEBEE';
  var bclr   = isRes ? '#2E7D32' : '#C62828';
  var btxt   = isRes ? 'AUTO-RESOLVED' : 'ESCALATED';
  var icdClr = c.icd10_valid ? '#2E7D32' : '#C62828';
  var icdTxt = c.icd10_valid ? '&#10003; Valid' : '&#10007; Invalid';
  var pc = {critical:'#C62828',high:'#E65100',medium:'#F9A825',low:'#2E7D32'}[(c.priority||'').toLowerCase()] || '#757575';
  var amt = '$'+(c.billed_amount||0).toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2});
  var div = document.createElement('div');
  div.style.cssText = 'background:#fff;border-radius:8px;border-left:5px solid '+border+';padding:16px 20px;box-shadow:0 1px 6px rgba(0,0,0,.07);margin-bottom:10px;opacity:0;transition:opacity .35s ease';
  div.innerHTML =
    '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">'
    +'<div><span style="font-weight:700;font-size:13px;color:#1565C0">'+(c.claim_id||'')+'</span>'
    +'<span style="margin-left:10px;font-size:12px;color:#424242">'+(c.patient_name||'')+'</span></div>'
    +'<span style="background:'+bbg+';color:'+bclr+';font-size:10px;font-weight:700;padding:3px 9px;border-radius:10px">'+btxt+'</span></div>'
    +'<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;font-size:11px;color:#424242;margin-bottom:8px">'
    +'<div><span style="color:#9E9E9E">Rejection: </span>'+(c.rejection_reason||'')+'</div>'
    +'<div><span style="color:#9E9E9E">ICD-10: </span><span style="color:'+icdClr+';font-weight:700">'+icdTxt+'</span></div>'
    +'<div><span style="color:#9E9E9E">Priority: </span><span style="color:'+pc+';font-weight:700">'+(c.priority||'').toUpperCase()+'</span></div>'
    +'</div>'
    +'<div style="font-size:11px"><span style="color:#9E9E9E">Billed: </span><strong>'+amt+'</strong>'
    +'<span style="color:#9E9E9E;margin-left:14px">Action: </span>'+(c.resolution_action||'')+'</div>';
  document.getElementById('billing-cards-grid').appendChild(div);
  setTimeout(function(){ div.style.opacity='1'; }, 40);
  var resolved = _billingOutcomes.filter(function(x){ return x.outcome==='auto_resolved'; });
  var rec  = resolved.reduce(function(s,x){ return s+(x.billed_amount||0); }, 0);
  var risk = _billingOutcomes.filter(function(x){ return x.outcome!=='auto_resolved'; }).reduce(function(s,x){ return s+(x.billed_amount||0); }, 0);
  document.getElementById('bb-total').textContent = _billingOutcomes.length;
  document.getElementById('bb-resolved').textContent = resolved.length;
  document.getElementById('bb-dollars-rec').textContent = '$'+Math.round(rec).toLocaleString('en-US');
  document.getElementById('bb-dollars-risk').textContent = '$'+Math.round(risk).toLocaleString('en-US');
}
function appendLine(text){
  if(text.startsWith('CLAIM_CARD:')){
    try{ addClaimCard(JSON.parse(text.slice(11))); }catch(e){}
    return;
  }
  if(text.startsWith('REPORT_INSIGHT:')){
    try{ addInsightCard(JSON.parse(text.slice(15))); }catch(e){}
    return;
  }
  if(text.startsWith('INTAKE_STARTING:')){
    updateBulkIntakeCardStarting(text.slice(16).trim());
    return;
  }
  if(text.startsWith('INTAKE_CARD:')){
    try{ updateBulkIntakeCardDone(JSON.parse(text.slice(12))); }catch(e){}
    return;
  }
  if(text.startsWith('ONBOARDING_STARTING:')){
    updateBulkCardStarting(text.slice(20).trim());
    return;
  }
  if(text.startsWith('ONBOARDING_CARD:')){
    try{ updateBulkCardDone(JSON.parse(text.slice(16))); }catch(e){}
    return;
  }
  const c = document.getElementById('console');
  const ph = c.querySelector('.ph');
  if(ph) ph.remove();
  const d = document.createElement('div');
  if(text.includes('[agent]   ->')) d.className='lt';
  else if(text.includes('[agent]   <-')) d.className='lr';
  else if(text.includes('iteration')) d.className='li';
  else if(text.includes('===') || text.includes('AGENT RESULT')) d.className='lf';
  else if(text.startsWith('[dashboard]') || text.startsWith('[intake]') || text.startsWith('[billing]') || text.startsWith('[demo]')) d.className='ld';
  d.textContent = text;
  c.appendChild(d);
  c.scrollTop = c.scrollHeight;
}
var _bulkIntakePatients = [];
function onBulkIntakeFileSelect(input){
  if(!input.files||!input.files[0]) return;
  var file = input.files[0];
  document.getElementById('bulk-intake-file-info').textContent = 'Uploading '+file.name+'...';
  var fd = new FormData();
  fd.append('file', file);
  fetch('/upload/bulk-intake',{method:'POST',body:fd})
    .then(function(r){ return r.json(); })
    .then(function(data){
      if(data.patients && data.patients.length > 0){
        _bulkIntakePatients = data.patients;
        var names = data.patients.map(function(p){ return (p.first_name||'')+' '+(p.last_name||''); }).join(', ');
        document.getElementById('bulk-intake-file-info').textContent =
          data.count+' patient'+(data.count>1?'s':'')+' found: '+names;
        document.getElementById('btn-bulk-intake').disabled = false;
      } else {
        document.getElementById('bulk-intake-file-info').textContent = 'No valid patient rows found in file.';
        document.getElementById('btn-bulk-intake').disabled = true;
      }
    })
    .catch(function(err){
      document.getElementById('bulk-intake-file-info').textContent = 'Upload failed: '+err;
    });
}
function clearBulkIntakeBoard(){
  document.getElementById('bulk-intake-cards-grid').innerHTML = '';
  document.getElementById('bulk-intake-board-wrap').style.display = 'none';
  document.getElementById('bki-total').textContent      = '0';
  document.getElementById('bki-complete').textContent   = '0';
  document.getElementById('bki-eligible').textContent   = '0';
  document.getElementById('bki-ineligible').textContent = '0';
}
function _updateBulkIntakeStats(){
  var cards    = document.querySelectorAll('.bki-card');
  var done     = document.querySelectorAll('.bki-card[data-status="success"]').length;
  var eligible = document.querySelectorAll('.bki-card[data-eligible="true"]').length;
  var inelig   = document.querySelectorAll('.bki-card[data-eligible="false"]').length;
  document.getElementById('bki-total').textContent      = cards.length;
  document.getElementById('bki-complete').textContent   = done;
  document.getElementById('bki-eligible').textContent   = eligible;
  document.getElementById('bki-ineligible').textContent = inelig;
}
function initBulkIntakeCard(patientName){
  document.getElementById('bulk-intake-board-wrap').style.display = 'block';
  var safeId = patientName.replace(/\s+/g,'-');
  var div = document.createElement('div');
  div.id = 'bki-'+safeId;
  div.className = 'bki-card';
  div.dataset.status = 'queued';
  div.style.cssText = 'background:#fff;border-radius:8px;border-left:5px solid #BDBDBD;padding:14px 20px;box-shadow:0 1px 6px rgba(0,0,0,.07);margin-bottom:8px;transition:border-color .25s';
  div.innerHTML =
    '<div style="display:flex;justify-content:space-between;align-items:center">'
    +'<span style="font-weight:700;font-size:13px;color:#1565C0">'+patientName+'</span>'
    +'<span id="bki-badge-'+safeId+'" style="background:#F5F5F5;color:#9E9E9E;font-size:10px;font-weight:700;padding:3px 9px;border-radius:10px">QUEUED</span>'
    +'</div>'
    +'<div id="bki-detail-'+safeId+'" style="font-size:12px;color:#9E9E9E;margin-top:6px">Waiting for batch slot...</div>';
  document.getElementById('bulk-intake-cards-grid').appendChild(div);
  _updateBulkIntakeStats();
}
function updateBulkIntakeCardStarting(patientName){
  var safeId = patientName.replace(/\s+/g,'-');
  var card = document.getElementById('bki-'+safeId);
  if(!card) return;
  card.dataset.status = 'running';
  card.style.borderLeftColor = '#E65100';
  var b = document.getElementById('bki-badge-'+safeId);
  b.style.background='#FFF3E0'; b.style.color='#E65100'; b.textContent='RUNNING';
  document.getElementById('bki-detail-'+safeId).textContent = 'Agent running...';
  _updateBulkIntakeStats();
}
function updateBulkIntakeCardDone(ins){
  var safeId = ins.patient_name.replace(/\s+/g,'-');
  var card = document.getElementById('bki-'+safeId);
  if(!card) return;
  var ok = ins.status==='success';
  card.dataset.status = ins.status;
  card.dataset.eligible = String(ins.is_eligible);
  var borderClr = ok ? (ins.is_eligible ? '#2E7D32' : '#C62828') : '#9E9E9E';
  card.style.borderLeftColor = borderClr;
  var b = document.getElementById('bki-badge-'+safeId);
  b.style.background = ok?'#E8F5E9':'#FFEBEE';
  b.style.color      = ok?'#2E7D32':'#C62828';
  b.textContent      = ok?'COMPLETE':'FAILED';
  var det = document.getElementById('bki-detail-'+safeId);
  if(ok){
    var eligBg  = ins.is_eligible ? '#E8F5E9' : '#FFEBEE';
    var eligClr = ins.is_eligible ? '#2E7D32' : '#C62828';
    var eligTxt = ins.is_eligible ? '&#10003; ELIGIBLE' : '&#10007; INELIGIBLE';
    var pdfLink = ins.pdf_key
      ? ' &nbsp;<a href="/intake/bulk-pdf/'+ins.pdf_key+'" target="_blank" style="font-size:10px;font-weight:700;color:#1565C0;text-decoration:none;padding:2px 8px;border-radius:10px;background:#E3F2FD">&#128196; View PDF</a>'
      : '';
    det.innerHTML =
      '<span style="color:#757575">'+(ins.insurance_id||'')+'</span>'
      +' &bull; <span style="color:#424242">'+(ins.plan_name||'')+'</span><br>'
      +'<span style="background:'+eligBg+';color:'+eligClr+';font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px">'+eligTxt+'</span>'
      +pdfLink;
  } else {
    det.innerHTML = '<span style="color:#C62828">'+(ins.error||'Unknown error')+'</span>';
  }
  _updateBulkIntakeStats();
}
function runBulkIntake(){
  if(!_bulkIntakePatients.length) return;
  clearBulkIntakeBoard();
  _bulkIntakePatients.forEach(function(p){
    initBulkIntakeCard(((p.first_name||'')+' '+(p.last_name||'')).trim());
  });
  setBadge('bulk-intake','running');
  document.getElementById('btn-bulk-intake').disabled = true;
  clearConsole();
  document.getElementById('console-label').textContent = 'Running: bulk intake ('+_bulkIntakePatients.length+' patients)...';
  var src = new EventSource('/run/bulk-intake');
  src.onmessage = function(e){
    var text = JSON.parse(e.data);
    if(text==='__DONE__'){
      src.close();
      setBadge('bulk-intake','success');
      document.getElementById('btn-bulk-intake').disabled = false;
      document.getElementById('console-label').textContent = 'Bulk intake — completed';
      loadAudit();
    } else { appendLine(text); }
  };
  src.onerror = function(){
    src.close();
    setBadge('bulk-intake','error');
    document.getElementById('btn-bulk-intake').disabled = false;
    appendLine('[ERROR] Bulk intake stream disconnected.');
    document.getElementById('console-label').textContent = 'Bulk intake — error';
  };
}
var _bulkEmployeeIds = [];
function onBulkFileSelect(input){
  if(!input.files||!input.files[0]) return;
  var file = input.files[0];
  document.getElementById('bulk-file-info').textContent = 'Uploading '+file.name+'...';
  var fd = new FormData();
  fd.append('file', file);
  fetch('/upload/bulk-onboarding',{method:'POST',body:fd})
    .then(function(r){ return r.json(); })
    .then(function(data){
      if(data.employee_ids && data.employee_ids.length > 0){
        _bulkEmployeeIds = data.employee_ids;
        document.getElementById('bulk-file-info').textContent =
          data.count+' employee'+(data.count>1?'s':'')+' found: '+data.employee_ids.join(', ');
        document.getElementById('btn-bulk').disabled = false;
      } else {
        document.getElementById('bulk-file-info').textContent = 'No valid EMP-XXXX IDs found in file.';
        document.getElementById('btn-bulk').disabled = true;
      }
    })
    .catch(function(err){
      document.getElementById('bulk-file-info').textContent = 'Upload failed: '+err;
    });
}
function clearBulkBoard(){
  document.getElementById('bulk-cards-grid').innerHTML = '';
  document.getElementById('bulk-board-wrap').style.display = 'none';
  document.getElementById('bko-total').textContent   = '0';
  document.getElementById('bko-complete').textContent = '0';
  document.getElementById('bko-running').textContent  = '0';
  document.getElementById('bko-failed').textContent   = '0';
}
function _updateBulkStats(){
  var cards  = document.querySelectorAll('.bko-card');
  var done   = document.querySelectorAll('.bko-card[data-status="success"]').length;
  var run    = document.querySelectorAll('.bko-card[data-status="running"]').length;
  var failed = document.querySelectorAll('.bko-card[data-status="failed"]').length;
  document.getElementById('bko-total').textContent    = cards.length;
  document.getElementById('bko-complete').textContent = done;
  document.getElementById('bko-running').textContent  = run;
  document.getElementById('bko-failed').textContent   = failed;
}
function initBulkCard(empId){
  document.getElementById('bulk-board-wrap').style.display = 'block';
  var div = document.createElement('div');
  div.id = 'bko-'+empId;
  div.className = 'bko-card';
  div.dataset.status = 'queued';
  div.style.cssText = 'background:#fff;border-radius:8px;border-left:5px solid #BDBDBD;padding:14px 20px;box-shadow:0 1px 6px rgba(0,0,0,.07);margin-bottom:8px;transition:border-color .25s';
  div.innerHTML =
    '<div style="display:flex;justify-content:space-between;align-items:center">'
    +'<span style="font-weight:700;font-size:13px;color:#1565C0">'+empId+'</span>'
    +'<span id="bko-badge-'+empId+'" style="background:#F5F5F5;color:#9E9E9E;font-size:10px;font-weight:700;padding:3px 9px;border-radius:10px">QUEUED</span>'
    +'</div>'
    +'<div id="bko-detail-'+empId+'" style="font-size:12px;color:#9E9E9E;margin-top:6px">Waiting for batch slot...</div>';
  document.getElementById('bulk-cards-grid').appendChild(div);
  _updateBulkStats();
}
function updateBulkCardStarting(empId){
  var card = document.getElementById('bko-'+empId);
  if(!card) return;
  card.dataset.status = 'running';
  card.style.borderLeftColor = '#E65100';
  var b = document.getElementById('bko-badge-'+empId);
  b.style.background='#FFF3E0'; b.style.color='#E65100'; b.textContent='RUNNING';
  document.getElementById('bko-detail-'+empId).textContent = 'Agent running...';
  _updateBulkStats();
}
function updateBulkCardDone(ins){
  var card = document.getElementById('bko-'+ins.employee_id);
  if(!card) return;
  var ok = ins.status==='success';
  card.dataset.status = ins.status;
  card.style.borderLeftColor = ok ? '#2E7D32' : '#C62828';
  var b = document.getElementById('bko-badge-'+ins.employee_id);
  b.style.background = ok?'#E8F5E9':'#FFEBEE';
  b.style.color      = ok?'#2E7D32':'#C62828';
  b.textContent      = ok?'COMPLETE':'FAILED';
  var det = document.getElementById('bko-detail-'+ins.employee_id);
  if(ok){
    var rCfg = {
      critical:{bg:'#FFEBEE',clr:'#C62828',icon:'&#9888;'},
      warning: {bg:'#FFF3E0',clr:'#E65100',icon:'&#9888;'},
      low:     {bg:'#E8F5E9',clr:'#2E7D32',icon:'&#10003;'},
      clear:   {bg:'#E8F5E9',clr:'#2E7D32',icon:'&#10003;'},
    }[ins.risk_level]||{bg:'#F5F5F5',clr:'#9E9E9E',icon:'&#9670;'};
    var rLbl = (ins.risk_level||'unknown').toUpperCase();
    var nLbl = (ins.notification_type||'standard').toUpperCase();
    var pdfBadge = '<a href="/onboarding/pdf/'+ins.employee_id+'?t='+Date.now()+'" target="_blank" style="font-size:10px;font-weight:700;color:#1565C0;text-decoration:none;padding:2px 8px;border-radius:10px;background:#E3F2FD;margin-left:6px">&#128196; View PDF</a>';
    det.innerHTML =
      '<span style="color:#212121;font-weight:600">'+(ins.employee_name||'')+'</span>'
      +' &bull; <span style="color:#757575">'+(ins.department||'')+'</span>'
      +' &bull; <span style="color:#9E9E9E;font-style:italic">'+(ins.position||'')+'</span><br>'
      +'<span style="background:'+rCfg.bg+';color:'+rCfg.clr+';font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px;margin-right:6px">'+rCfg.icon+'&nbsp;RISK: '+rLbl+'</span>'
      +'<span style="background:#E3F2FD;color:#1565C0;font-size:10px;font-weight:700;padding:2px 8px;border-radius:10px">'+nLbl+'</span>'
      +pdfBadge;
  } else {
    det.innerHTML = '<span style="color:#C62828">'+( ins.error||'Unknown error')+'</span>';
  }
  _updateBulkStats();
}
function runBulk(){
  if(!_bulkEmployeeIds.length) return;
  clearBulkBoard();
  _bulkEmployeeIds.forEach(function(id){ initBulkCard(id); });
  setBadge('bulk','running');
  document.getElementById('btn-bulk').disabled = true;
  clearConsole();
  document.getElementById('console-label').textContent = 'Running: bulk onboarding ('+_bulkEmployeeIds.length+' employees)...';
  var src = new EventSource('/run/bulk-onboarding?ids='+encodeURIComponent(_bulkEmployeeIds.join(',')));
  src.onmessage = function(e){
    var text = JSON.parse(e.data);
    if(text==='__DONE__'){
      src.close();
      setBadge('bulk','success');
      document.getElementById('btn-bulk').disabled = false;
      document.getElementById('console-label').textContent = 'Bulk onboarding — completed';
      loadAudit();
    } else { appendLine(text); }
  };
  src.onerror = function(){
    src.close();
    setBadge('bulk','error');
    document.getElementById('btn-bulk').disabled = false;
    appendLine('[ERROR] Bulk onboarding stream disconnected.');
    document.getElementById('console-label').textContent = 'Bulk onboarding — error';
  };
}
function clearConsole(){
  document.getElementById('console').innerHTML = '<span class="ph">Output will stream here in real time when you click Run above.</span>';
}
function run(automation){
  let url = '/run/'+automation;
  if(automation==='onboarding'){
    url += '?employee_id='+encodeURIComponent(document.getElementById('emp-id').value||'EMP-0023');
  } else if(automation==='report'){
    url += '?mode='+document.getElementById('report-mode').value;
  }
  if(automation==='billing') clearBillingBoard();
  if(automation==='report')  clearReportBoard();
  if(automation==='onboarding'){
    document.getElementById('link-onboarding-pdf').style.display = 'none';
  }
  setBadge(automation,'running');
  setBtn(automation,true);
  clearConsole();
  document.getElementById('console-label').textContent = 'Running: '+automation+'...';
  const src = new EventSource(url);
  src.onmessage = function(e){
    const text = JSON.parse(e.data);
    if(text==='__DONE__'){
      src.close();
      setBadge(automation,'success');
      setBtn(automation,false);
      showResult(automation);
      document.getElementById('console-label').textContent = automation+' — completed';
      loadAudit();
    } else {
      appendLine(text);
    }
  };
  src.onerror = function(){
    src.close();
    setBadge(automation,'error');
    setBtn(automation,false);
    appendLine('[ERROR] Stream disconnected.');
    document.getElementById('console-label').textContent = automation+' — error';
  };
}
function loadAudit(){
  fetch('/audit').then(r=>r.json()).then(data=>{
    const body = document.getElementById('audit-body');
    if(!data.entries||data.entries.length===0){
      body.innerHTML='<div class="empty">No entries yet. Run an automation first.</div>';
      return;
    }
    body.innerHTML = data.entries.slice(0,40).map(e=>{
      const d = new Date(e.timestamp);
      const ts = d.toLocaleDateString()+' '+d.toLocaleTimeString();
      return '<div class="audit-row">'
        +'<span class="at">'+ts+'</span>'
        +'<span class="aa">'+e.automation+'</span>'
        +'<span class="ac">'+e.action+'</span>'
        +'<span class="ae">'+(e.entity_id||'&mdash;')+'</span>'
        +'</div>';
    }).join('');
  });
}
let _apiCheckCount = 0;
function checkApis(){
  fetch('/api/status').then(r=>r.json()).then(d=>{
    const allReady = Object.values(d).every(v=>v==='ready');
    const dot = document.getElementById('api-dot');
    const lbl = document.getElementById('api-label');
    dot.className = 'dot'+(allReady?' ready':'');
    if(allReady){
      lbl.textContent = 'Mock APIs ready';
    } else if(++_apiCheckCount < 30){
      lbl.textContent = 'APIs starting... ('+_apiCheckCount+'/30)';
      setTimeout(checkApis,2000);
    } else {
      lbl.textContent = 'APIs offline — restart dashboard';
    }
  }).catch(()=>{ if(++_apiCheckCount<30) setTimeout(checkApis,2000); });
}
checkApis();
loadAudit();
setInterval(loadAudit,15000);
</script>
</body>
</html>"""


if __name__ == "__main__":
    uvicorn.run("dashboard:app", host="0.0.0.0", port=8000, reload=False)
